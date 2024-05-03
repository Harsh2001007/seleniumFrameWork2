import openpyxl
from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws1 = wb.create_sheet("sheet-1")
ws2 = wb.create_sheet("sheet-2")
ws3 = wb.create_sheet("sheet-3")

print(wb.sheetnames)
ws["A4"] = "hello"


wb.save("testing.xlsx")
